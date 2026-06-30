import ctypes
import os, sys
import pandas as pd
import Planilhas
from openpyxl import load_workbook
from pandastable import Table
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename


def armazenaImagem(diretorioAtual):
    try:
        diretorioTemporario = sys._MEIPASS
    except Exception:
        diretorioTemporario = os.path.abspath('.')
    
    return os.path.join(diretorioTemporario, diretorioAtual)

dadosProcessadosMinibio = False
nomesDuplicadosProcessado = None
planilhasMescladasProcessado = None

def carregaPlanilhas(tipo):
    diretorio = askopenfilename(filetypes=[('Excel files', '*.xlsx *.xls')])

    if not diretorio:
        return
    
    def testaPlanilhasLideresMensaisCarregadas():
        if Planilhas.planilhaMinibio is not None and Planilhas.planilhaErgon is not None:
            buttonValoresDiferentes.config(state='normal', bg=corAzul)
            buttonNomesDuplicados.config(state='normal', bg=corAzul)

    def validaTipoPlanilha(tipo):
        tipos = {'minibio': ['minibio'], 'ergon': ['ergon']}

        nomeArquivo = diretorio.lower()

        return any(palavra in nomeArquivo for palavra in tipos.get(tipo, []))

    if not validaTipoPlanilha(tipo):
        labelMensagem['text'] = 'Planilha errada.'
        labelMensagem.after(5000, lambda: labelMensagem.config(text=''))

        return

    if tipo == 'minibio':
        Planilhas.planilhaMinibio = pd.read_excel(diretorio)

        global planilhaMinibioAtiva, planilhaMinibioFormatada

        planilhaMinibioFormatada = load_workbook(diretorio)
        planilhaMinibioAtiva = planilhaMinibioFormatada.active

        labelMensagem['text'] = 'Planilha MINIBIO carregada.'

        testaPlanilhasLideresMensaisCarregadas()
    elif tipo == 'ergon':
        Planilhas.planilhaErgon = pd.read_excel(diretorio)

        labelMensagem['text'] = 'Planilha ERGON carregada.'

        testaPlanilhasLideresMensaisCarregadas()

def mostraPlanilha(planilha, titulo, fechaJanelaDadosDuplicados=None):
    if planilha is None or planilha.empty:
        labelMensagem['text'] = 'Não há dados para exibir.'
    else:
        novaJanela = Toplevel(janela)
        novaJanela.title(titulo)
        novaJanela.geometry('1100x500')
        novaJanela.resizable(False, False)

        framePlanilha = Frame(novaJanela)
        framePlanilha.pack(fill=BOTH, expand=1)

        tabela = Table(framePlanilha, dataframe=planilha)
        tabela.show()

        if fechaJanelaDadosDuplicados:
            novaJanela.protocol("WM_DELETE_WINDOW", lambda:fechaJanelaDadosDuplicados(novaJanela))

def garantirProcessamentoLideres():
    global dadosProcessadosMinibio, nomesDuplicadosProcessado, planilhasMescladasProcessado
    if not dadosProcessadosMinibio:
        nomesDuplicadosProcessado, planilhasMescladasProcessado, planilhaAlterada = Planilhas.processaPlanilhasLideresMensais()

        if nomesDuplicadosProcessado is None:
            return False

        for linha in planilhaMinibioAtiva.iter_rows():
            for coluna in linha:
                coluna.value = None
        
        for indiceColuna, coluna in enumerate(planilhaAlterada.columns, start=1):
            planilhaMinibioAtiva.cell(row=1, column=indiceColuna, value=coluna)

        for indiceLinha, linha in planilhaAlterada.iterrows():
            for indiceColuna, coluna in enumerate(linha, start=1):
                planilhaMinibioAtiva.cell(row=indiceLinha + 2, column=indiceColuna, value=coluna)

        for i in range(planilhaMinibioAtiva.max_row, 1, -1):
            if planilhaMinibioAtiva.cell(row=i, column=1).value is None:
                planilhaMinibioAtiva.delete_rows(i)
            
        planilhaMinibioFormatada.save('Planilha Minibio - eliminados registros de ex líderes.xlsx')
        dadosProcessadosMinibio = True
    return True

def mostraValoresDiferentes():
    if not garantirProcessamentoLideres():
        labelMensagem['text'] = 'Carregue as planilhas primeiro!'
        return
    
    colunasExibir = ['NOME', 'ORGAO_ENTIDADE_ERGON', 'ORGAO_ENTIDADE_MINIBIO', 'NOMESETOR_ERGON', 'NOMESETOR_MINIBIO', 'SIGLA']
    if 'CPF' in planilhasMescladasProcessado.columns:
        colunasExibir.insert(1, 'CPF')
    if 'REFERENCIA_MINIBIO' in planilhasMescladasProcessado.columns and 'REFERENCIA_ERGON' in planilhasMescladasProcessado.columns:
        colunasExibir.extend(['REFERENCIA_MINIBIO', 'REFERENCIA_ERGON'])
        
    valoresDiferentes = planilhasMescladasProcessado.loc[planilhasMescladasProcessado['IGUAIS'] == False, colunasExibir]
    
    def confirmarSalvamento(janelaPopup):
        resposta = messagebox.askyesno("Salvar Alterações", "Deseja salvar esses valores diferentes na planilha Histórico?")
        
        if resposta:  
            diretorioHistorico = askopenfilename(title="Selecione a planilha Histórico", filetypes=[('Excel files', '*.xlsx *.xls')])
            if not diretorioHistorico:
                labelMensagem['text'] = 'Ação cancelada. Planilha Histórico não selecionada.'
                janelaPopup.destroy()
                return
                
            try:
                wb = load_workbook(diretorioHistorico)
                ws = wb.active

                cabecalho = [cell.value for cell in ws[1]]

                for _, linha in valoresDiferentes.iterrows():
                    novaLinha = []
                    for coluna in cabecalho:
                        novaLinha.append(linha.get(coluna, None))
                    ws.append(novaLinha)

                wb.save(diretorioHistorico)
                labelMensagem['text'] = 'Histórico atualizado com sucesso!'
            except Exception as e:
                labelMensagem['text'] = f'Erro ao salvar: {str(e)}'
        else:
            labelMensagem['text'] = 'Ação cancelada. Dados não foram salvos.'
        
        janelaPopup.destroy()  

    fechaJanela = confirmarSalvamento
    
    mostraPlanilha(valoresDiferentes, 'Valores Diferentes', fechaJanelaDadosDuplicados=fechaJanela)

def mostraNomesDuplicados():
    if not garantirProcessamentoLideres():
        labelMensagem['text'] = 'Carregue as planilhas primeiro!'
        return
    # colunasDuplicadas = ['NOME', 'INICIO_LOTACAO', 'NOMESETOR', 'ORGAO_ENTIDADE']
    colunasDuplicadas = ['NOME', 'NOMESETOR', 'ORGAO_ENTIDADE']
    if 'CPF' in nomesDuplicadosProcessado.columns:
        colunasDuplicadas.insert(1, 'CPF')
    mostraPlanilha(nomesDuplicadosProcessado[colunasDuplicadas], 'Nomes Duplicados')

def resetaTudo():
    global dadosProcessadosMinibio
    dadosProcessadosMinibio = False
    
    buttonPlanilhaMinibio.config(state='normal', bg=corAzul)
    buttonPlanilhaErgon.config(state='normal', bg=corAzul)
    
    buttonValoresDiferentes.config(state='disabled', bg=corCinza)
    buttonNomesDuplicados.config(state='disabled', bg=corCinza)

    Planilhas.planilhaMinibio = None
    Planilhas.planilhaErgon = None

    labelMensagem['text'] = ''

#Interface
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("ProcessadorPlanilhas") #Forçar o logo a aparecer em toda a interface do sistema

corAzul = '#0085B3'
corAzulEscuro = '#094A75'
corBorda = '#D6E6F1'
corCard = '#FFFFFF'
corCinza = '#B5B5B5'
corFundo = '#ECF1F4'
corTexto = '#1D1D1B'

estiloBotao = {'bg': corAzul, 'fg': 'white', 'activebackground': corAzulEscuro, 'activeforeground': 'white', 'font': ('Segoe UI', 10, 'bold'), 'relief': 'flat', 'bd': 0, 'width': 22, 'cursor': 'hand2', 'pady': 6}

infosLideresMensais = """Esta ferramenta faz:

1. Elimina registros da planilha minibio que não estejam na planilha Ergon;
2. Salva uma nova planilha sem esses registros eliminados;
3. Exibe registros das duas planilhas que estejam com valores diferentes para ORGAO_ENTIDADE, NOME_SETOR e REFERENCIA;
4. Exibe registros de nomes duplicados da planilha minibio;
5. Salva em uma planilha da sua escolha os registros com valores diferentes.

OBS: A planilha minibio precisa conter a palavra 'minibio' no nome e a planilha do Ergon precisa conter a palavra 'ergon' no nome"""

janela = Tk()
janela.iconbitmap(armazenaImagem('iconeInterface.ico'))
janela.geometry('1100x650')
janela.resizable(False, False)
janela.title('Processador de Planilhas Lideres Cariocas')
janela.configure(bg=corFundo)

labelTitulo = Label(janela, text='Processador de Planilhas\nLíderes Cariocas', bg=corFundo, fg=corAzulEscuro, font=('Segoe UI', 26, 'bold'))
labelTitulo.pack(pady=(25, 20))


frameInstrucoes = Frame(janela, bg=corCard, highlightbackground=corBorda, highlightthickness=1)
frameInstrucoes.pack(pady=(0, 20), padx=50, fill='x')

frameTituloInst = Frame(frameInstrucoes, bg=corCard)
frameTituloInst.pack(fill='x', padx=20, pady=(15, 5))
Label(frameTituloInst, text='⚙️', bg=corCard, fg=corAzul, font=('Segoe UI Emoji', 20)).pack(side='left')
Label(frameTituloInst, text='Instruções', bg=corCard, fg=corAzulEscuro, font=('Segoe UI', 14, 'bold')).pack(side='left', padx=10)

Label(frameInstrucoes, text=infosLideresMensais, bg=corCard, fg=corTexto, justify='left', font=('Segoe UI', 10)).pack(anchor='w', padx=25, pady=(0, 15))


frameBotoesUpload = Frame(janela, bg=corFundo)
frameBotoesUpload.pack(pady=10)

buttonPlanilhaMinibio = Button(frameBotoesUpload, text='Planilha Minibio', command=lambda: carregaPlanilhas('minibio'), **estiloBotao)
buttonPlanilhaMinibio.pack(side='left', padx=10)

buttonPlanilhaErgon = Button(frameBotoesUpload, text='Planilha Ergon', command=lambda: carregaPlanilhas('ergon'), **estiloBotao)
buttonPlanilhaErgon.pack(side='left', padx=10)


frameBotoesAcao = Frame(janela, bg=corFundo)
frameBotoesAcao.pack(pady=20)

buttonValoresDiferentes = Button(frameBotoesAcao, text='Valores Diferentes', command=mostraValoresDiferentes, **estiloBotao)
buttonValoresDiferentes.config(state='disabled', bg=corCinza)
buttonValoresDiferentes.pack(side='left', padx=10)

buttonNomesDuplicados = Button(frameBotoesAcao, text='Nomes Duplicados', command=mostraNomesDuplicados, **estiloBotao)
buttonNomesDuplicados.config(state='disabled', bg=corCinza)
buttonNomesDuplicados.pack(side='left', padx=10)


labelMensagem = Label(janela, bg=corFundo, fg=corAzulEscuro, font=('Segoe UI', 10, 'bold'))
labelMensagem.pack(pady=10)

buttonRefresh = Button(janela, text='↻', command=lambda: resetaTudo(), fg=corAzul, activeforeground=corAzulEscuro, font=('Segoe UI', 25, 'bold'), relief='flat', bd=0, width=3, height=1, cursor='hand2')
buttonRefresh.place(relx=1.0, rely=1.0, x=-25, y=-20, anchor='se')

janela.mainloop()